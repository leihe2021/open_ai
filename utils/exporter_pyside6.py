#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
数据导出模块 (PySide6版本)
支持将预约记录导出为Excel或CSV格式
"""

import csv
import os
from datetime import datetime
from typing import List, Tuple, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class DataExporter:
    """数据导出类 (PySide6版本)"""

    # 表头配置
    HEADERS = [
        "ID",
        "院区",
        "血制品大类",
        "血制品亚类",
        "血型",
        "数量",
        "预约时间"
    ]

    # 列宽配置（Excel）
    COLUMN_WIDTHS = {
        "ID": 8,
        "院区": 15,
        "血制品大类": 15,
        "血制品亚类": 15,
        "血型": 10,
        "数量": 10,
        "预约时间": 20
    }

    def __init__(self, parent_window=None):
        """初始化导出器"""
        self.parent = parent_window

    def export_to_excel(self, data: List[Tuple], output_file: str) -> bool:
        """
        导出数据到Excel文件

        Args:
            data: 数据列表，每行是一个元组
            output_file: 输出文件路径

        Returns:
            bool: 是否成功
        """
        if not HAS_OPENPYXL:
            self._show_error("错误", "未安装 openpyxl 库，无法导出Excel文件")
            return False

        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "血制品预约记录"

            # 写入表头
            for col_num, header in enumerate(self.HEADERS, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)

            # 设置列宽
            for col_num, header in enumerate(self.HEADERS, 1):
                col_letter = self._get_column_letter(col_num)
                ws.column_dimensions[col_letter].width = self.COLUMN_WIDTHS.get(header, 12)

            # 写入数据
            for row_num, row_data in enumerate(data, 2):
                for col_num, cell_value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                    cell.alignment = Alignment(horizontal='center' if col_num != 2 else 'left', vertical='center')

            # 保存文件
            wb.save(output_file)
            self._show_info("成功", f"数据已成功导出到：\n{output_file}")
            return True

        except Exception as e:
            self._show_error("导出失败", f"导出Excel时发生错误：\n{str(e)}")
            return False

    def export_to_csv(self, data: List[Tuple], output_file: str) -> bool:
        """
        导出数据到CSV文件

        Args:
            data: 数据列表，每行是一个元组
            output_file: 输出文件路径

        Returns:
            bool: 是否成功
        """
        try:
            with open(output_file, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile)
                # 写入表头
                writer.writerow(self.HEADERS)
                # 写入数据
                writer.writerows(data)

            self._show_info("成功", f"数据已成功导出到：\n{output_file}")
            return True

        except Exception as e:
            self._show_error("导出失败", f"导出CSV时发生错误：\n{str(e)}")
            return False

    def export_data(self, data: List[Tuple], file_format: str = "xlsx") -> bool:
        """
        导出数据（根据格式自动选择）

        Args:
            data: 数据列表
            file_format: 文件格式，"xlsx" 或 "csv"

        Returns:
            bool: 是否成功
        """
        # 获取文件路径
        initial_file = self._get_output_path(file_format)
        if not initial_file:
            return False  # 用户取消

        try:
            if file_format.lower() == "xlsx":
                return self.export_to_excel(data, initial_file)
            elif file_format.lower() == "csv":
                return self.export_to_csv(data, initial_file)
            else:
                self._show_error("错误", f"不支持的格式：{file_format}")
                return False
        except Exception as e:
            self._show_error("导出失败", f"发生未知错误：\n{str(e)}")
            return False

    def _get_output_path(self, file_format: str) -> Optional[str]:
        """
        获取输出文件路径

        Args:
            file_format: 文件格式

        Returns:
            str: 文件路径或None（用户取消）
        """
        # 使用PySide6的文件对话框
        from PySide6.QtWidgets import QFileDialog, QApplication

        # 设置文件扩展名
        if file_format.lower() == "xlsx":
            filter_str = "Excel文件 (*.xlsx)"
            defextension = ".xlsx"
        elif file_format.lower() == "csv":
            filter_str = "CSV文件 (*.csv)"
            defextension = ".csv"
        else:
            return None

        # 生成默认文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"血制品预约记录_{timestamp}{defextension}"

        # 显示保存对话框
        dialog = QFileDialog(self.parent, "保存导出文件", filename, filter_str)
        dialog.setDefaultSuffix(defextension[1:])  # 去掉点号

        if dialog.exec():
            filepath = dialog.selectedFiles()[0]
            return filepath

        return None

    def _get_column_letter(self, col_num: int) -> str:
        """将列号转换为Excel列字母（如：1->A, 2->B）"""
        from openpyxl.utils import get_column_letter
        return get_column_letter(col_num)

    def _show_info(self, title: str, message: str):
        """显示信息对话框"""
        if self.parent:
            try:
                from PySide6.QtWidgets import QMessageBox
                QMessageBox.information(self.parent, title, message)
            except:
                print(f"[INFO] {title}: {message}")
        else:
            print(f"[INFO] {title}: {message}")

    def _show_error(self, title: str, message: str):
        """显示错误对话框"""
        if self.parent:
            try:
                from PySide6.QtWidgets import QMessageBox
                QMessageBox.critical(self.parent, title, message)
            except:
                print(f"[ERROR] {title}: {message}")
        else:
            print(f"[ERROR] {title}: {message}")
